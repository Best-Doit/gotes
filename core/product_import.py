from io import BytesIO
import unicodedata
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException


MAX_PRODUCT_ROWS = 5000
MAX_REPORTED_ERRORS = 100

HEADER_ALIASES = {
    "codigo": "code",
    "code": "code",
    "nombre": "name",
    "producto": "name",
    "name": "name",
    "categoria": "category",
    "category": "category",
}
HEADER_LABELS = {
    "code": "Código",
    "name": "Nombre",
    "category": "Categoría",
}
FIELD_LIMITS = {"code": 60, "name": 200, "category": 100}


class InvalidProductWorkbook(Exception):
    pass


def _plain_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_header(value):
    text = _plain_text(value).lower()
    text = "".join(
        character for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )
    return " ".join(text.split())


def build_product_template():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Productos"
    worksheet.append(("Código", "Nombre", "Categoría"))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:C1"
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 26
    worksheet.column_dimensions["A"].number_format = "@"

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="286BC3")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 24

    instructions = workbook.create_sheet("Instrucciones")
    instructions.append(("Importación de productos GOTES",))
    instructions.append(("1. Completa las columnas Código, Nombre y Categoría en la hoja Productos.",))
    instructions.append(("2. No cambies los encabezados ni repitas códigos dentro del archivo.",))
    instructions.append(("3. Un código existente en tu empresa actualizará su nombre y categoría.",))
    instructions.append((f"4. Se permiten hasta {MAX_PRODUCT_ROWS} productos por archivo.",))
    instructions.column_dimensions["A"].width = 92
    instructions["A1"].font = Font(bold=True, size=14)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def parse_product_workbook(uploaded_file):
    uploaded_file.seek(0)
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        raise InvalidProductWorkbook("El archivo no es un Excel .xlsx válido o está dañado.") from error

    try:
        worksheet = workbook.active
        row_iterator = worksheet.iter_rows(values_only=True)
        headers = next(row_iterator, None)
        if not headers:
            raise InvalidProductWorkbook("El archivo está vacío.")

        columns = {}
        duplicate_headers = []
        for index, value in enumerate(headers):
            field = HEADER_ALIASES.get(_normalized_header(value))
            if not field:
                continue
            if field in columns:
                duplicate_headers.append(HEADER_LABELS[field])
            else:
                columns[field] = index

        missing = [HEADER_LABELS[field] for field in HEADER_LABELS if field not in columns]
        if duplicate_headers:
            raise InvalidProductWorkbook(f"Hay encabezados repetidos: {', '.join(duplicate_headers)}.")
        if missing:
            raise InvalidProductWorkbook(f"Faltan las columnas obligatorias: {', '.join(missing)}.")

        products = []
        errors = []
        errors_truncated = False
        seen_codes = set()
        data_rows = 0

        def report(line, message):
            nonlocal errors_truncated
            if len(errors) < MAX_REPORTED_ERRORS:
                errors.append({"line": line, "message": message})
            else:
                errors_truncated = True

        for line, values in enumerate(row_iterator, start=2):
            if not any(_plain_text(value) for value in values):
                continue
            data_rows += 1
            if data_rows > MAX_PRODUCT_ROWS:
                report(line, f"El archivo supera el máximo de {MAX_PRODUCT_ROWS} productos.")
                break

            row = {}
            for field, index in columns.items():
                row[field] = _plain_text(values[index] if index < len(values) else None)

            for field, label in HEADER_LABELS.items():
                if not row[field]:
                    report(line, f"{label} es obligatorio.")
                elif len(row[field]) > FIELD_LIMITS[field]:
                    report(line, f"{label} supera el máximo de {FIELD_LIMITS[field]} caracteres.")

            if row["code"]:
                if row["code"] in seen_codes:
                    report(line, f"El código {row['code']} está repetido en el archivo.")
                seen_codes.add(row["code"])

            products.append({"line": line, **row})

        if not products and not errors:
            report(2, "No se encontraron productos para importar.")
        return products, errors, errors_truncated
    finally:
        workbook.close()
