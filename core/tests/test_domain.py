import tempfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.core import mail
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.db import close_old_connections
from django.test import TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from core.models import (
    AuditLog,
    Branch,
    Company,
    EmailOutbox,
    Evidence,
    Incident,
    IncidentDifference,
    Notification,
    Product,
    ReceiptItem,
    Sequence,
    Transfer,
    TransferItem,
    User,
)
from core.services import (
    cancel_transfer,
    close_transfer,
    confirm_receipt,
    dispatch_transfer,
    get_or_start_receipt,
    prepare_transfer,
    resolve_incident,
    save_commercial_registration,
    visible_transfers,
)


TEST_MEDIA = tempfile.mkdtemp(prefix="gotes-test-media-")


def product_excel(rows, headers=("Código", "Nombre", "Categoría")):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@override_settings(
    MEDIA_ROOT=TEST_MEDIA,
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    GOTES_PUBLIC_URL="https://gotes.example.test",
)
class DomainTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.company_a = Company.objects.create(code="empresa-a", name="Empresa A")
        cls.company_b = Company.objects.create(code="empresa-b", name="Empresa B")
        cls.a_origin = Branch.objects.create(company=cls.company_a, code="AO", name="A Origen")
        cls.a_destination = Branch.objects.create(company=cls.company_a, code="AD", name="A Destino")
        cls.b_origin = Branch.objects.create(company=cls.company_b, code="BO", name="B Origen")
        cls.b_destination = Branch.objects.create(company=cls.company_b, code="BD", name="B Destino")
        cls.a_product = Product.objects.create(company=cls.company_a, code="A-1", name="Producto A", category="General")
        cls.a_extra = Product.objects.create(company=cls.company_a, code="A-2", name="Producto extra", category="General")
        cls.b_product = Product.objects.create(company=cls.company_b, code="B-1", name="Producto B", category="General")
        cls.a_manager_origin = User.objects.create_user(
            username="a-manager-origin", password="test-pass-123", company=cls.company_a,
            branch=cls.a_origin, role=User.Role.MANAGER, email="origin@example.test",
        )
        cls.a_manager_destination = User.objects.create_user(
            username="a-manager-destination", password="test-pass-123", company=cls.company_a,
            branch=cls.a_destination, role=User.Role.MANAGER, email="destination@example.test",
        )
        cls.a_restricted_manager = User.objects.create_user(
            username="a-restricted-manager", password="test-pass-123", company=cls.company_a,
            branch=cls.a_origin, role=User.Role.MANAGER, allow_dispatch=False,
        )
        cls.a_admin = User.objects.create_user(
            username="a-admin", password="test-pass-123", company=cls.company_a,
            role=User.Role.COMPANY_ADMIN,
        )
        cls.a_reconciler = User.objects.create_user(
            username="a-reconciler", password="test-pass-123", company=cls.company_a,
            role=User.Role.RECONCILER, email="reconciler@example.test",
        )
        cls.b_manager = User.objects.create_user(
            username="b-manager", password="test-pass-123", company=cls.company_b,
            branch=cls.b_origin, role=User.Role.MANAGER,
        )

    def make_transfer(self, *, company=None, origin=None, destination=None, user=None, product=None, quantity="10"):
        company = company or self.company_a
        origin = origin or self.a_origin
        destination = destination or self.a_destination
        user = user or self.a_manager_origin
        product = product or self.a_product
        transfer = Transfer.objects.create(company=company, origin=origin, destination=destination, created_by=user)
        TransferItem.objects.create(transfer=transfer, product=product, quantity_sent=Decimal(quantity))
        return transfer

    def add_evidence(self, transfer, evidence_type, user):
        return Evidence.objects.create(
            transfer=transfer,
            type=evidence_type,
            file=SimpleUploadedFile("evidencia.jpg", b"\xff\xd8\xfffake-image-content", content_type="image/jpeg"),
            uploaded_by=user,
        )

    def dispatch(self, transfer):
        prepare_transfer(transfer, self.a_manager_origin)
        self.add_evidence(transfer, Evidence.Type.DISPATCH, self.a_manager_origin)
        return dispatch_transfer(transfer, self.a_manager_origin)

    def receive_exact(self, transfer):
        receipt = get_or_start_receipt(transfer, self.a_manager_destination)
        self.add_evidence(transfer, Evidence.Type.RECEIPT, self.a_manager_destination)
        confirm_receipt(transfer, self.a_manager_destination)
        transfer.refresh_from_db()
        return receipt

    def test_correlatives_are_scoped_by_company_and_year(self):
        first_a = self.make_transfer()
        second_a = self.make_transfer()
        first_b = self.make_transfer(
            company=self.company_b,
            origin=self.b_origin,
            destination=self.b_destination,
            user=self.b_manager,
            product=self.b_product,
        )
        self.assertTrue(first_a.code.endswith("000001"))
        self.assertTrue(second_a.code.endswith("000002"))
        self.assertTrue(first_b.code.endswith("000001"))
        self.assertNotEqual(first_a.uuid, first_b.uuid)

    def test_business_roles_are_consolidated(self):
        self.assertEqual(
            [value for value, _label in User.Role.choices],
            ["SUPERADMIN", "COMPANY_ADMIN", "MANAGER", "RECONCILER", "AUDITOR"],
        )
        self.assertEqual(
            [label for _value, label in User.Role.choices],
            ["Superusuario", "Administrador de empresa", "Encargado", "Conciliador comercial", "Auditor"],
        )
        self.assertEqual(User._meta.get_field("role").default, "")
        self.assertTrue(User._meta.get_field("role").blank)

    def test_destination_cannot_see_draft_but_sees_dispatched(self):
        transfer = self.make_transfer()
        self.assertFalse(visible_transfers(self.a_manager_destination).filter(pk=transfer.pk).exists())
        self.dispatch(transfer)
        self.assertTrue(visible_transfers(self.a_manager_destination).filter(pk=transfer.pk).exists())

    def test_exact_happy_path_requires_photos_and_manual_close(self):
        transfer = self.make_transfer()
        prepare_transfer(transfer, self.a_manager_origin)
        with self.assertRaises(ValidationError):
            dispatch_transfer(transfer, self.a_manager_origin)
        self.add_evidence(transfer, Evidence.Type.DISPATCH, self.a_manager_origin)
        dispatch_transfer(transfer, self.a_manager_origin)
        get_or_start_receipt(transfer, self.a_manager_destination)
        with self.assertRaises(ValidationError):
            confirm_receipt(transfer, self.a_manager_destination)
        self.add_evidence(transfer, Evidence.Type.RECEIPT, self.a_manager_destination)
        confirm_receipt(transfer, self.a_manager_destination)
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.RECEIVED)
        close_transfer(transfer, self.a_manager_destination)
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.CLOSED)

    def test_manager_operational_permission_can_be_restricted(self):
        transfer = self.make_transfer(user=self.a_restricted_manager)
        prepare_transfer(transfer, self.a_restricted_manager)
        self.add_evidence(transfer, Evidence.Type.DISPATCH, self.a_restricted_manager)
        with self.assertRaises(PermissionDenied):
            dispatch_transfer(transfer, self.a_restricted_manager)
        self.a_restricted_manager.allow_dispatch = True
        self.a_restricted_manager.save(update_fields=("allow_dispatch",))
        dispatch_transfer(transfer, self.a_restricted_manager)
        receipt = get_or_start_receipt(transfer, self.a_manager_destination)
        self.add_evidence(transfer, Evidence.Type.RECEIPT, self.a_manager_destination)
        confirm_receipt(transfer, self.a_manager_destination)
        receipt.refresh_from_db()
        self.assertEqual(receipt.status, receipt.Status.CONFIRMED)

    def test_differences_are_grouped_resolved_then_closed(self):
        transfer = self.make_transfer()
        self.dispatch(transfer)
        receipt = get_or_start_receipt(transfer, self.a_manager_destination)
        expected = receipt.items.get(is_unexpected=False)
        expected.quantity_received = Decimal("8")
        expected.is_damaged = True
        expected.observation = "Envase dañado"
        expected.save()
        ReceiptItem.objects.create(
            receipt=receipt,
            product=self.a_extra,
            quantity_received=Decimal("2"),
            is_unexpected=True,
            observation="No figuraba en el envío",
        )
        self.add_evidence(transfer, Evidence.Type.RECEIPT, self.a_manager_destination)
        confirm_receipt(transfer, self.a_manager_destination)
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.RECEIVED_DIFFERENCES)
        incident = Incident.objects.get(transfer=transfer)
        self.assertEqual(incident.differences.count(), 3)
        self.assertSetEqual(
            set(incident.differences.values_list("type", flat=True)),
            {IncidentDifference.Type.MISSING, IncidentDifference.Type.DAMAGED, IncidentDifference.Type.UNEXPECTED},
        )
        with self.assertRaises(ValidationError):
            close_transfer(transfer, self.a_manager_destination)
        resolve_incident(incident, self.a_manager_origin, Incident.ResolutionType.ACCEPTED, "Diferencia verificada.")
        close_transfer(transfer, self.a_manager_destination)
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.CLOSED)

    def test_commercial_registration_is_separate_and_correction_needs_reason(self):
        transfer = self.make_transfer()
        self.dispatch(transfer)
        self.receive_exact(transfer)
        self.assertTrue(
            self.a_reconciler.notifications.filter(message__contains="pendiente de conciliación comercial").exists()
        )
        self.assertFalse(
            self.a_admin.notifications.filter(message__contains="pendiente de conciliación comercial").exists()
        )
        received_deliveries = EmailOutbox.objects.filter(
            transfer=transfer,
            event=EmailOutbox.Event.TRANSFER_RECEIVED,
        )
        self.assertSetEqual(
            set(received_deliveries.values_list("recipient_email", flat=True)),
            {"origin@example.test", "destination@example.test", "reconciler@example.test"},
        )
        reconciler_email = received_deliveries.get(recipient=self.a_reconciler)
        self.assertIn("fue recepcionado", reconciler_email.subject)
        self.assertIn("requiere conciliación comercial", reconciler_email.body)
        self.assertIn("Abrir conciliación", reconciler_email.body)
        manager_emails = received_deliveries.filter(recipient__role=User.Role.MANAGER)
        self.assertEqual(manager_emails.count(), 2)
        self.assertTrue(all("Abrir seguimiento" in delivery.body for delivery in manager_emails))
        registration = save_commercial_registration(transfer, self.a_reconciler, {
            "external_reference": "ERP-100", "external_date": date.today(), "notes": "",
        })
        self.assertEqual(registration.external_reference, "ERP-100")
        self.assertEqual(
            EmailOutbox.objects.filter(transfer=transfer, event=EmailOutbox.Event.TRANSFER_RECONCILED).count(),
            2,
        )
        with self.assertRaises(PermissionDenied):
            save_commercial_registration(transfer, self.a_admin, {
                "external_reference": "ERP-ADMIN", "external_date": date.today(), "notes": "",
            })
        with self.assertRaises(ValidationError):
            save_commercial_registration(transfer, self.a_reconciler, {
                "external_reference": "ERP-101", "external_date": date.today(), "notes": "",
            })
        save_commercial_registration(transfer, self.a_reconciler, {
            "external_reference": "ERP-101", "external_date": date.today(), "notes": "Corregido",
        }, reason="Referencia digitada incorrectamente")
        transfer.refresh_from_db()
        self.assertEqual(transfer.commercial_registration.external_reference, "ERP-101")
        self.assertEqual(
            EmailOutbox.objects.filter(transfer=transfer, event=EmailOutbox.Event.TRANSFER_RECONCILED).count(),
            2,
        )

    def test_dispatch_queues_and_worker_sends_email_to_both_branch_managers(self):
        transfer = self.make_transfer()
        self.dispatch(transfer)

        deliveries = EmailOutbox.objects.filter(
            transfer=transfer,
            event=EmailOutbox.Event.TRANSFER_DISPATCHED,
        )
        self.assertSetEqual(
            set(deliveries.values_list("recipient_email", flat=True)),
            {"origin@example.test", "destination@example.test", "reconciler@example.test"},
        )
        self.assertTrue(all("https://gotes.example.test" in delivery.body for delivery in deliveries))

        call_command("send_notification_emails")

        self.assertEqual(len(mail.outbox), 3)
        self.assertEqual(deliveries.filter(status=EmailOutbox.Status.SENT).count(), 3)
        self.assertTrue(all(message.subject == f"[GOTES] {transfer.code} fue despachado" for message in mail.outbox))

    def test_items_are_immutable_after_dispatch(self):
        transfer = self.make_transfer()
        self.dispatch(transfer)
        item = transfer.items.get()
        item.quantity_sent = Decimal("1")
        with self.assertRaises(ValidationError):
            item.full_clean()

    def test_closed_transfer_cannot_be_cancelled(self):
        transfer = self.make_transfer()
        self.dispatch(transfer)
        self.receive_exact(transfer)
        close_transfer(transfer, self.a_manager_destination)
        with self.assertRaises(ValidationError):
            cancel_transfer(transfer, self.a_manager_destination, "No debería poder")


@override_settings(MEDIA_ROOT=TEST_MEDIA)
class TenantIsolationViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.company_a = Company.objects.create(code="isolation-a", name="Isolation A")
        cls.company_b = Company.objects.create(code="isolation-b", name="Isolation B")
        cls.a_origin = Branch.objects.create(company=cls.company_a, code="AO", name="A Origen")
        cls.a_destination = Branch.objects.create(company=cls.company_a, code="AD", name="A Destino")
        cls.b_origin = Branch.objects.create(company=cls.company_b, code="BO", name="B Origen")
        cls.b_destination = Branch.objects.create(company=cls.company_b, code="BD", name="B Destino")
        cls.a_product = Product.objects.create(company=cls.company_a, code="A", name="Producto A", category="General")
        cls.b_product = Product.objects.create(company=cls.company_b, code="B", name="Producto B", category="General")
        cls.a_admin = User.objects.create_user(username="isolation-admin-a", password="x", company=cls.company_a, role=User.Role.COMPANY_ADMIN)
        cls.a_reconciler = User.objects.create_user(username="isolation-reconciler-a", password="x", company=cls.company_a, role=User.Role.RECONCILER)
        cls.a_auditor = User.objects.create_user(username="isolation-auditor-a", password="x", company=cls.company_a, role=User.Role.AUDITOR)
        cls.a_manager = User.objects.create_user(username="isolation-manager-a", password="x", company=cls.company_a, branch=cls.a_origin, role=User.Role.MANAGER)
        cls.a_destination_manager = User.objects.create_user(username="isolation-manager-destination", password="x", company=cls.company_a, branch=cls.a_destination, role=User.Role.MANAGER)
        cls.b_manager = User.objects.create_user(username="isolation-manager-b", password="x", company=cls.company_b, branch=cls.b_origin, role=User.Role.MANAGER)
        cls.b_reconciler = User.objects.create_user(username="isolation-reconciler-b", password="x", company=cls.company_b, role=User.Role.RECONCILER)
        cls.b_auditor = User.objects.create_user(username="isolation-auditor-b", password="x", company=cls.company_b, role=User.Role.AUDITOR)

    def make_transfer(self, company, origin, destination, user, product):
        transfer = Transfer.objects.create(company=company, origin=origin, destination=destination, created_by=user)
        TransferItem.objects.create(transfer=transfer, product=product, quantity_sent=1)
        return transfer

    def add_evidence(self, transfer, user):
        return Evidence.objects.create(
            transfer=transfer,
            type=Evidence.Type.PREPARATION,
            file=SimpleUploadedFile("isolation.jpg", b"\xff\xd8\xfffile", content_type="image/jpeg"),
            uploaded_by=user,
        )

    def setUp(self):
        self.a_transfer = self.make_transfer(
            self.company_a, self.a_origin, self.a_destination, self.a_manager, self.a_product,
        )
        self.b_transfer = self.make_transfer(
            self.company_b, self.b_origin, self.b_destination, self.b_manager, self.b_product,
        )
        self.b_evidence = self.add_evidence(self.b_transfer, self.b_manager)

    def test_other_company_uuid_returns_404(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("transfer_detail", args=(self.b_transfer.uuid,)))
        self.assertEqual(response.status_code, 404)

    def test_other_company_evidence_returns_404(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("evidence_download", args=(self.b_evidence.uuid,)))
        self.assertEqual(response.status_code, 404)

    def test_csv_never_contains_other_company(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("transfer_export_csv"))
        content = response.content.decode("utf-8")
        self.assertIn(self.a_transfer.code, content)
        self.assertIn(self.a_origin.name, content)
        self.assertNotIn(self.b_origin.name, content)

    def test_company_admin_cannot_use_django_admin(self):
        self.client.force_login(self.a_admin)
        response = self.client.get("/django-admin/")
        self.assertEqual(response.status_code, 302)

    def test_company_admin_cannot_edit_other_company_catalog(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("manage_product_edit", args=(self.b_product.pk,)))
        self.assertEqual(response.status_code, 404)

    def test_company_admin_cannot_edit_other_company_assignment(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("manage_assignment_edit", args=(self.b_manager.pk,)))
        self.assertEqual(response.status_code, 404)

    def test_operational_pages_render(self):
        self.client.force_login(self.a_manager)
        dashboard = self.client.get(reverse("dashboard"))
        self.assertEqual(dashboard.status_code, 200)
        self.assertContains(dashboard, 'class="stats ')
        self.assertContains(dashboard, "border-y border-base-300 bg-transparent shadow-none")
        self.assertContains(dashboard, 'class="border-t border-base-300 pt-3"')
        self.assertNotContains(dashboard, 'class="rounded-box bg-base-200 p-3"')
        self.assertContains(dashboard, 'class="table premium-table table-sm"')
        self.assertContains(dashboard, "border-spacing: 0;")
        self.assertContains(dashboard, "box-shadow: 0 10px 30px")
        self.assertContains(dashboard, "font-variant-numeric: tabular-nums")
        self.assertContains(dashboard, "padding-block: .55rem")
        self.assertNotContains(dashboard, "border-spacing: .25rem .35rem")
        self.assertContains(dashboard, "--color-primary: #286bc3")
        self.assertContains(dashboard, "--color-base-100: #ffffff")
        self.assertContains(dashboard, "--color-base-200: #f0f4f8")
        self.assertNotContains(dashboard, "radial-gradient(circle at 92% 4%")
        self.assertContains(dashboard, ".btn-static-warning:is(:hover, :focus-visible, :active)")
        self.assertContains(dashboard, ".btn.btn-soft:not(.sidebar-new):is(:hover, :focus-visible, :active)")
        self.assertContains(dashboard, ".btn.btn-error.btn-soft:not(.sidebar-new)")
        self.assertContains(dashboard, ".input:is(:focus, :focus-within, :focus-visible)")
        self.assertContains(dashboard, "box-shadow: 0 0 0 1px color-mix(in oklab, var(--color-primary) 28%, transparent)")
        self.assertNotContains(dashboard, "box-shadow: 0 0 0 3px color-mix(in oklab, var(--color-primary) 7%, transparent)")
        self.assertContains(dashboard, ".checkbox:focus-visible")
        self.assertContains(dashboard, 'class="sidebar-new btn btn-primary btn-soft')
        self.assertContains(dashboard, 'class="btn btn-primary btn-soft btn-sm px-4"')
        self.assertContains(dashboard, 'class="badge badge-neutral badge-outline"')
        self.assertContains(dashboard, 'class="btn btn-soft btn-sm flex-1"')
        self.assertContains(dashboard, 'class="btn btn-error btn-soft btn-sm w-full"')
        self.assertNotContains(dashboard, "background: color-mix(in oklab, var(--color-primary) 11%, transparent)")
        self.assertContains(dashboard, 'class="drawer lg:drawer-open"')
        self.assertContains(dashboard, 'id="app-navigation"')
        self.assertContains(dashboard, 'class="navbar sticky')
        self.assertContains(dashboard, 'class="navbar sticky top-2')
        self.assertNotContains(dashboard, "rounded-selector bg-primary/10 p-1 pl-2")
        self.assertContains(dashboard, "border border-base-300 bg-base-100 px-5 shadow-sm lg:flex")
        self.assertNotContains(dashboard, "bg-base-100/95")
        self.assertContains(dashboard, 'class="app-sidebar m-2 mr-0')
        self.assertContains(dashboard, "flex-1 px-2 pb-2 pt-2")
        self.assertContains(dashboard, "menu-active")
        self.assertContains(dashboard, "menu-active bg-primary/10 font-bold text-primary")
        self.assertNotContains(dashboard, "menu-active bg-base-300")
        self.assertNotContains(dashboard, "Administración")
        self.assertContains(dashboard, 'id="global-product-suggestions"')
        self.assertContains(dashboard, '<option value="A">Producto A</option>', html=True)
        self.assertNotContains(dashboard, "Producto B")
        transfer_list = self.client.get(reverse("transfer_list"))
        self.assertEqual(transfer_list.status_code, 200)
        self.assertContains(transfer_list, "Buscar traspasos")
        self.assertContains(transfer_list, "Más filtros")
        self.assertContains(transfer_list, "Movimiento")
        self.assertContains(transfer_list, "Conciliación comercial")
        self.assertContains(transfer_list, "Responsable")
        self.assertContains(transfer_list, "Ver detalle")
        self.assertContains(transfer_list, 'name="q"')
        self.assertNotContains(transfer_list, "<th>Creación</th>")
        self.assertContains(transfer_list, 'Aplicar filtros</button>')
        self.assertNotContains(transfer_list, 'btn btn-primary btn-sm w-full px-4')
        self.assertContains(transfer_list, "Nuevo")
        create_page = self.client.get(reverse("transfer_create"))
        self.assertEqual(create_page.status_code, 200)
        self.assertContains(create_page, 'class="card card-xs border border-base-300 bg-base-100 shadow-sm"')
        self.assertContains(create_page, "md:grid-cols-[minmax(9rem,.55fr)_minmax(12rem,1fr)_minmax(16rem,1.45fr)]")
        self.assertContains(create_page, 'class="textarea textarea-sm h-8 min-h-8 w-full resize-none py-1.5"')
        self.assertContains(create_page, 'rows="1"')
        self.assertContains(create_page, 'class="table premium-table transfer-items-table table-sm min-w-[46rem]"')
        self.assertContains(create_page, ".premium-table.transfer-items-table tbody tr:hover td")
        self.assertContains(create_page, ".premium-table.transfer-items-table tbody td")
        self.assertContains(create_page, '.transfer-items-table input[type="number"]::-webkit-inner-spin-button')
        self.assertContains(create_page, '.transfer-items-table input[type="search"]::-webkit-search-cancel-button')
        self.assertContains(create_page, "-moz-appearance: textfield")
        self.assertContains(create_page, '<tbody id="item-rows">')
        self.assertContains(create_page, '<th class="w-28">Código</th>')
        self.assertContains(create_page, '<th class="w-[34%]">Nombre</th>')
        self.assertContains(create_page, '<template id="empty-item-row">')
        self.assertContains(create_page, 'x-data="productPicker()"')
        self.assertContains(create_page, 'placeholder="Escribe nombre o código"')
        self.assertContains(create_page, 'x-text="product.name"')
        self.assertContains(create_page, "x-text=\"selectedCode || '—'\"")
        self.assertContains(create_page, 'class="input input-sm w-full"')
        self.assertContains(create_page, 'this.selectedCode = product.code;')
        self.assertContains(create_page, 'this.nativeSelect.value = product.value;')
        self.assertContains(create_page, 'class="hidden" x-ref="nativeSelect"')
        self.assertNotContains(create_page, "querySelector('select').options")
        self.assertContains(create_page, "Agregar fila</button>")
        self.assertContains(create_page, 'class="btn btn-info btn-soft btn-sm"')
        self.assertNotContains(create_page, 'class="divide-y divide-base-300" id="item-rows"')
        self.assertNotContains(create_page, 'class="rounded-box bg-base-200 p-2"')
        reports = self.client.get(reverse("reports"))
        self.assertEqual(reports.status_code, 200)
        self.assertContains(reports, 'class="btn btn-info btn-soft btn-sm"')
        self.assertContains(reports, "Resumen operativo")
        self.assertContains(reports, "Mapa de carga del flujo")
        self.assertContains(reports, "Control del proceso")
        self.assertContains(reports, "Carga operativa por sucursal")
        self.assertContains(reports, "Puntos de atención")
        self.assertContains(reports, 'name="date_from"')
        self.assertContains(reports, 'class="radial-progress')
        self.assertContains(reports, 'class="steps steps-horizontal')
        self.assertContains(reports, '<section class="card')
        self.assertNotContains(reports, '<table')
        self.assertNotContains(reports, 'class="rounded-box border border-base-300 bg-base-100 p-3')
        detail = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)), {"flow": "1"})
        self.assertEqual(detail.status_code, 200)
        self.assertContains(detail, 'class="btn action-button btn-primary btn-sm"')
        self.assertContains(detail, ".btn.action-button.btn-success:is(:hover, :focus-visible, :active)")
        self.assertContains(detail, ".btn.action-button.btn-error:is(:hover, :focus-visible, :active)")
        self.assertContains(detail, 'id="transfer-flow-modal"')
        self.assertContains(detail, "Ver más detalle")
        self.assertContains(detail, 'aria-controls="transfer-flow-modal"')
        self.assertContains(detail, "max-w-2xl")
        self.assertContains(detail, 'id="evidence-modal"')
        self.assertContains(detail, 'id="cancel-transfer-modal"')
        self.assertContains(detail, 'class="steps steps-horizontal')
        self.assertContains(detail, '<th>Código</th><th>Producto</th><th>Categoría</th>')
        self.assertNotContains(detail, ">Acciones</h2>")
        self.assertNotContains(detail, 'lg:grid-cols-[minmax(0,2fr)_minmax(19rem,0.75fr)]')
        self.assertNotContains(detail, 'class="collapse collapse-arrow')
        self.assertContains(detail, "Historial")
        self.assertContains(detail, "Ver solo la tabla")
        self.assertNotContains(detail, "✕")
        self.assertNotContains(detail, "↗")
        self.assertContains(detail, 'const openFlow = () => { if (flow && !flow.open) flow.showModal(); };')
        self.assertNotContains(detail, "La conciliación estará disponible después de confirmar la recepción.")
        self.assertContains(detail, "Anular movimiento")

    def test_framework_messages_render_in_global_modal(self):
        self.client.force_login(self.a_manager)
        response = self.client.post(
            reverse("transfer_prepare", args=(self.a_transfer.uuid,)),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="global-message-modal"')
        self.assertContains(response, "Mensaje del sistema")
        self.assertContains(response, "Traspaso preparado.")
        self.assertContains(response, 'modal.showModal()')
        self.assertContains(response, 'class="modal-backdrop bg-neutral/20 backdrop-blur-[1px]"')
        self.assertNotContains(response, 'class="alert mb-4')

    def test_saving_draft_opens_operational_flow_modal(self):
        self.client.force_login(self.a_manager)
        item = self.a_transfer.items.get()
        response = self.client.post(
            reverse("transfer_edit", args=(self.a_transfer.uuid,)),
            {
                "destination": self.a_destination.pk,
                "notes": "Listo para preparar",
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "1",
                "items-MIN_NUM_FORMS": "0",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-id": item.pk,
                "items-0-product": self.a_product.pk,
                "items-0-quantity_sent": item.quantity_sent,
                "items-0-send_note": "",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.redirect_chain[-1][0].endswith("?flow=1&saved=1"))
        self.assertContains(response, 'id="transfer-flow-modal"')
        self.assertContains(response, "Borrador guardado correctamente")
        self.assertContains(response, "Confirmar preparación")
        self.assertNotContains(response, 'id="global-message-modal"')

    def test_dispatch_screen_explains_evidence_requirement_and_allows_exit(self):
        prepare_transfer(self.a_transfer, self.a_manager)
        self.client.force_login(self.a_manager)

        blocked_page = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)))
        self.assertContains(blocked_page, ">Adjuntar evidencia</button>")
        self.assertContains(blocked_page, "Necesitas una evidencia de tipo")
        self.assertContains(blocked_page, 'id="evidence-modal"')
        self.assertContains(blocked_page, '<option value="DISPATCH" selected>Salida</option>', html=True)
        self.assertNotContains(blocked_page, ">Confirmar salida</button>")

        evidence = self.add_evidence(self.a_transfer, self.a_manager)
        evidence.type = Evidence.Type.DISPATCH
        evidence.save(update_fields=("type",))
        ready_page = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)))
        self.assertContains(ready_page, ">Confirmar salida</button>")
        self.assertContains(ready_page, 'id="dispatch-confirm-modal"')

        response = self.client.post(
            reverse("transfer_dispatch", args=(self.a_transfer.uuid,)),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.a_transfer.refresh_from_db()
        self.assertEqual(self.a_transfer.status, Transfer.Status.DISPATCHED)
        self.assertContains(response, "Salida confirmada.")

    def test_global_product_search_stays_inside_company(self):
        self.client.force_login(self.a_manager)
        response = self.client.get(reverse("transfer_list"), {"q": self.a_product.name})
        self.assertContains(response, self.a_transfer.code)
        self.assertNotContains(response, self.b_origin.name)
        self.assertNotContains(response, str(self.b_transfer.uuid))

    def test_receipt_page_renders_for_destination(self):
        prepare_transfer(self.a_transfer, self.a_manager)
        evidence = self.add_evidence(self.a_transfer, self.a_manager)
        evidence.type = Evidence.Type.DISPATCH
        evidence.save(update_fields=("type",))
        dispatch_transfer(self.a_transfer, self.a_manager)
        self.client.force_login(self.a_destination_manager)
        response = self.client.get(reverse("receipt_edit", args=(self.a_transfer.uuid,)))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Productos por recibir")
        self.assertContains(response, "Continuar recepción")
        self.assertContains(response, 'id="receipt-flow-modal"')
        self.assertContains(response, 'class="steps steps-horizontal')
        self.assertContains(response, 'x-data="{ step: 1 }"')
        self.assertContains(response, 'id="receipt-draft-form"')
        self.assertContains(response, 'id="receipt-evidence-form"')
        self.assertContains(response, 'id="receipt-confirm-form"')
        self.assertContains(response, 'name="_return_to" value="receipt"')
        self.assertContains(response, "Guardar y continuar")
        self.assertContains(response, "Adjuntar y continuar")
        self.assertContains(response, "Confirmar recepción")
        self.assertContains(response, 'class="table premium-table table-sm"')

        receipt_item = self.a_transfer.receipt.items.get(is_unexpected=False)
        saved = self.client.post(reverse("receipt_edit", args=(self.a_transfer.uuid,)), {
            "expected-TOTAL_FORMS": "1",
            "expected-INITIAL_FORMS": "1",
            "expected-MIN_NUM_FORMS": "0",
            "expected-MAX_NUM_FORMS": "1000",
            "expected-0-id": receipt_item.pk,
            "expected-0-quantity_received": "1.000",
            "expected-0-observation": "",
            "unexpected-TOTAL_FORMS": "1",
            "unexpected-INITIAL_FORMS": "0",
            "unexpected-MIN_NUM_FORMS": "0",
            "unexpected-MAX_NUM_FORMS": "1000",
            "unexpected-0-id": "",
            "unexpected-0-product": "",
            "unexpected-0-quantity_received": "",
            "unexpected-0-observation": "",
        })
        self.assertEqual(saved.status_code, 302)
        self.assertTrue(saved["Location"].endswith("?flow=1&step=3&saved=1"))

        evidence_saved = self.client.post(reverse("upload_evidence", args=(self.a_transfer.uuid,)), {
            "type": Evidence.Type.RECEIPT,
            "file": SimpleUploadedFile("receipt.jpg", b"\xff\xd8\xffreceipt-proof", content_type="image/jpeg"),
            "description": "Recepción física",
            "_return_to": "receipt",
        })
        self.assertEqual(evidence_saved.status_code, 302)
        self.assertTrue(evidence_saved["Location"].endswith("?flow=1&step=4&evidence=1"))

    def test_company_management_pages_render(self):
        self.client.force_login(self.a_admin)
        branches = self.client.get(reverse("manage_branches"))
        self.assertEqual(branches.status_code, 200)
        self.assertContains(branches, "Administración")
        self.assertContains(branches, "menu-active")
        products = self.client.get(reverse("manage_products"))
        self.assertEqual(products.status_code, 200)
        self.assertContains(products, "Importar Excel")
        self.assertContains(products, f'{reverse("manage_products")}?importar=1')
        self.assertContains(products, 'id="product-import-modal"')
        self.assertContains(products, f'action="{reverse("product_import")}"')
        product_import_modal = self.client.get(reverse("manage_products"), {"importar": "1"})
        self.assertContains(product_import_modal, "document.getElementById('product-import-modal').showModal()")
        users = self.client.get(reverse("manage_users"), {"nuevo": "1"})
        self.assertEqual(users.status_code, 200)
        self.assertContains(users, self.a_admin.username)
        self.assertContains(users, "Administrador de empresa")
        self.assertContains(users, "Conciliador comercial")
        self.assertContains(users, "Solo lectura")
        self.assertContains(users, "flex max-h-[calc(100dvh-1rem)]")
        self.assertContains(users, "min-h-0 flex-1 overflow-y-auto overscroll-contain")
        self.assertContains(users, "grid shrink-0 grid-cols-2")
        self.assertContains(users, 'class="table premium-table table-sm"')
        self.assertContains(users, 'class="modal-box premium-modal')
        self.assertContains(users, 'class="modal-close btn btn-circle btn-ghost btn-sm shrink-0"')
        self.assertContains(users, ".modal-close:is(:hover, :active)")
        self.assertNotContains(users, 'aria-label="Cerrar">✕</button>')
        self.assertContains(users, 'class="btn btn-outline btn-sm">Cancelar</a>')
        self.assertContains(users, 'class="modal-backdrop bg-neutral/20 backdrop-blur-[1px]"')
        self.assertContains(users, "grid gap-y-2")
        self.assertContains(users, "premium-modal .fieldset")
        self.assertContains(users, 'name="username"')
        self.assertContains(users, 'name="first_name"')
        self.assertContains(users, 'name="email"')
        self.assertContains(users, 'name="password1"')
        self.assertNotContains(users, 'name="role"')
        self.assertNotContains(users, 'name="branch"')
        self.assertNotContains(users, 'name="phone"')
        self.assertNotContains(users, 'name="allow_dispatch"')
        self.assertNotContains(users, "rounded-box bg-base-200 p-3")
        self.assertNotContains(users, "sm:grid-cols-2")
        self.assertNotContains(users, "modal-action sticky")

        assignments = self.client.get(reverse("manage_assignments"), {"nuevo": "1"})
        self.assertEqual(assignments.status_code, 200)
        self.assertContains(assignments, "Asignaciones de usuarios")
        self.assertContains(assignments, "Nueva asignación")
        self.assertContains(assignments, "Encargado")
        self.assertContains(assignments, "Conciliador comercial")
        self.assertContains(assignments, "Auditor")
        self.assertContains(assignments, 'name="user"')
        self.assertContains(assignments, 'name="role"')
        self.assertContains(assignments, 'name="branch"')
        self.assertContains(assignments, 'class="table premium-table table-sm"')
        self.assertContains(assignments, 'id="assignment-form-modal"')
        self.assertContains(assignments, "menu-active")
        self.assertEqual(self.client.get(reverse("audit_list")).status_code, 200)

    def test_product_excel_template_has_required_columns(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("product_import_template"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        content = b"".join(response.streaming_content)
        workbook = load_workbook(BytesIO(content), read_only=True)
        self.assertEqual(
            [cell.value for cell in next(workbook["Productos"].iter_rows(max_row=1))],
            ["Código", "Nombre", "Categoría"],
        )
        self.assertIn("Instrucciones", workbook.sheetnames)
        workbook.close()

    def test_company_admin_imports_and_updates_products_from_excel(self):
        other_company_product = Product.objects.create(
            company=self.company_b,
            code=self.a_product.code,
            name="Producto ajeno",
            category="Otra empresa",
        )
        upload = SimpleUploadedFile(
            "productos.xlsx",
            product_excel([
                (self.a_product.code, "Producto A actualizado", "Actualizados"),
                ("A-NEW", "Producto nuevo", "General"),
            ]),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(self.a_admin)
        response = self.client.post(reverse("product_import"), {"file": upload})
        self.assertRedirects(response, reverse("manage_products"))

        self.a_product.refresh_from_db()
        other_company_product.refresh_from_db()
        self.assertEqual(self.a_product.name, "Producto A actualizado")
        self.assertEqual(self.a_product.category, "Actualizados")
        self.assertEqual(other_company_product.name, "Producto ajeno")
        self.assertTrue(Product.objects.filter(company=self.company_a, code="A-NEW", name="Producto nuevo").exists())
        self.assertEqual(
            AuditLog.objects.filter(company=self.company_a, user=self.a_admin, action="IMPORT_PRODUCT").count(),
            2,
        )

    def test_invalid_excel_import_is_atomic_and_reports_rows(self):
        upload = SimpleUploadedFile(
            "productos.xlsx",
            product_excel([
                ("DUP-1", "Primero", "General"),
                ("DUP-1", "Repetido", "General"),
                ("", "Sin código", "General"),
            ]),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(self.a_admin)
        response = self.client.post(reverse("product_import"), {"file": upload})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "está repetido en el archivo")
        self.assertContains(response, "Código es obligatorio")
        self.assertContains(response, "No se importó ningún producto")
        self.assertContains(response, 'id="product-import-modal"')
        self.assertContains(response, "document.getElementById('product-import-modal').showModal()")
        self.assertFalse(Product.objects.filter(company=self.company_a, code="DUP-1").exists())

    def test_only_company_admin_can_import_products(self):
        upload = SimpleUploadedFile(
            "productos.xlsx",
            product_excel([("NO-AUTH", "No autorizado", "General")]),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(self.a_auditor)
        self.assertEqual(self.client.get(reverse("product_import")).status_code, 403)
        self.assertEqual(self.client.get(reverse("product_import_template")).status_code, 403)
        self.assertEqual(self.client.post(reverse("product_import"), {"file": upload}).status_code, 403)
        self.assertFalse(Product.objects.filter(company=self.company_a, code="NO-AUTH").exists())

    def test_company_admin_creates_minimal_user_pending_assignment(self):
        self.client.force_login(self.a_admin)
        response = self.client.post(reverse("manage_users"), {
            "username": "new-user",
            "first_name": "Rosa",
            "email": "rosa@example.com",
            "branch": self.a_origin.pk,
            "role": User.Role.RECONCILER,
            "password1": "Maple!9384-Quartz!River",
        })
        self.assertRedirects(
            response,
            reverse("manage_users"),
            msg_prefix=str(response.context["form"].errors) if response.status_code == 200 else "",
        )
        user = User.objects.get(username="new-user")
        self.assertEqual(user.first_name, "Rosa")
        self.assertEqual(user.email, "rosa@example.com")
        self.assertEqual(user.role, "")
        self.assertIsNone(user.branch_id)
        self.assertIsNone(user.allow_dispatch)
        self.assertTrue(user.check_password("Maple!9384-Quartz!River"))

    def test_company_admin_assigns_role_and_branch_from_assignments(self):
        user = User.objects.create_user(
            username="pending-manager",
            password="Cobalt!5729-Lake!Stone",
            company=self.company_a,
            role="",
        )
        self.client.force_login(self.a_admin)
        response = self.client.post(reverse("manage_assignments"), {
            "user": user.pk,
            "branch": self.a_origin.pk,
            "role": User.Role.MANAGER,
        })
        self.assertRedirects(response, reverse("manage_assignments"))
        user.refresh_from_db()
        self.assertEqual(user.role, User.Role.MANAGER)
        self.assertEqual(user.branch, self.a_origin)
        self.assertTrue(user.is_operational)
        self.assertTrue(AuditLog.objects.filter(user=self.a_admin, action="MANAGE_ASSIGNMENT").exists())

    def test_assignment_requires_branch_only_for_manager(self):
        user = User.objects.create_user(
            username="pending-auditor",
            password="Quartz!5729-Lake!Stone",
            company=self.company_a,
            role="",
        )
        self.client.force_login(self.a_admin)
        missing_branch = self.client.post(reverse("manage_assignments"), {
            "user": user.pk,
            "role": User.Role.MANAGER,
            "branch": "",
        })
        self.assertEqual(missing_branch.status_code, 200)
        self.assertContains(missing_branch, "Selecciona la sucursal del encargado")
        user.refresh_from_db()
        self.assertEqual(user.role, "")

        assigned = self.client.post(reverse("manage_assignments"), {
            "user": user.pk,
            "role": User.Role.AUDITOR,
            "branch": self.a_origin.pk,
        })
        self.assertRedirects(assigned, reverse("manage_assignments"))
        user.refresh_from_db()
        self.assertEqual(user.role, User.Role.AUDITOR)
        self.assertIsNone(user.branch_id)

    def test_reconciler_has_company_scope_but_no_operational_or_admin_access(self):
        self.client.force_login(self.a_reconciler)
        response = self.client.get(reverse("transfer_list"))
        self.assertContains(response, self.a_transfer.code)
        self.assertNotContains(response, self.b_origin.name)
        self.assertNotContains(response, str(self.b_transfer.uuid))
        self.assertEqual(self.client.get(reverse("transfer_create")).status_code, 403)
        self.assertEqual(self.client.get(reverse("manage_users")).status_code, 403)
        self.assertEqual(self.client.get(reverse("manage_assignments")).status_code, 403)
        self.assertEqual(self.client.get(reverse("audit_list")).status_code, 200)

    def test_reconciler_registers_reference_and_manager_can_consult_it(self):
        prepare_transfer(self.a_transfer, self.a_manager)
        dispatch_evidence = self.add_evidence(self.a_transfer, self.a_manager)
        dispatch_evidence.type = Evidence.Type.DISPATCH
        dispatch_evidence.save(update_fields=("type",))
        dispatch_transfer(self.a_transfer, self.a_manager)
        get_or_start_receipt(self.a_transfer, self.a_destination_manager)
        receipt_evidence = self.add_evidence(self.a_transfer, self.a_destination_manager)
        receipt_evidence.type = Evidence.Type.RECEIPT
        receipt_evidence.save(update_fields=("type",))
        confirm_receipt(self.a_transfer, self.a_destination_manager)

        self.client.force_login(self.a_reconciler)
        detail = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)))
        self.assertContains(detail, "Registrar conciliación")
        self.assertContains(detail, f'value="{date.today().isoformat()}"')
        self.assertContains(detail, 'type="date"')
        self.assertNotContains(detail, "Motivo de corrección")
        self.assertNotContains(detail, 'id="commercial-correction-modal"')
        self.assertContains(detail, "Agregar observación opcional")
        response = self.client.post(reverse("commercial_register", args=(self.a_transfer.uuid,)), {
            "external_reference": "SIS-TR-908",
            "external_date": date.today().isoformat(),
            "notes": "Registrado al cierre del turno",
            "correction_reason": "",
        })
        self.assertRedirects(response, f"{reverse('transfer_detail', args=(self.a_transfer.uuid,))}?flow=1")
        self.a_transfer.refresh_from_db()
        self.assertEqual(self.a_transfer.commercial_registration.registered_by, self.a_reconciler)

        correction_detail = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)))
        self.assertContains(correction_detail, "Corregir registro")
        self.assertContains(correction_detail, 'id="commercial-correction-modal"')
        self.assertContains(correction_detail, "openTransferSubflow('commercial-correction-modal')")
        self.assertContains(correction_detail, "Guardar corrección")
        self.assertContains(correction_detail, "Motivo de corrección")
        self.assertContains(correction_detail, "Obligatorio para conservar el historial de la corrección")

        self.client.force_login(self.a_destination_manager)
        detail = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)))
        self.assertContains(detail, "SIS-TR-908")
        self.assertContains(detail, "Registrado")
        self.assertNotContains(detail, "Guardar conciliación")

        self.client.force_login(self.a_admin)
        denied = self.client.post(reverse("commercial_register", args=(self.a_transfer.uuid,)), {
            "external_reference": "SIS-TR-909",
            "external_date": date.today().isoformat(),
            "notes": "",
            "correction_reason": "Cambio no autorizado",
        })
        self.assertEqual(denied.status_code, 403)
        self.a_transfer.refresh_from_db()
        self.assertEqual(self.a_transfer.commercial_registration.external_reference, "SIS-TR-908")

    def test_reconciler_cannot_see_or_modify_another_company(self):
        self.client.force_login(self.a_reconciler)
        self.assertEqual(
            self.client.get(reverse("transfer_detail", args=(self.b_transfer.uuid,))).status_code,
            404,
        )
        self.assertEqual(
            self.client.post(reverse("commercial_register", args=(self.b_transfer.uuid,)), {
                "external_reference": "FORBIDDEN",
                "external_date": date.today().isoformat(),
                "notes": "",
                "correction_reason": "",
            }).status_code,
            404,
        )

    def test_auditor_sees_all_company_information_in_read_only_mode(self):
        self.client.force_login(self.a_auditor)

        dashboard = self.client.get(reverse("dashboard"))
        self.assertContains(dashboard, "Modo auditor")
        transfers = self.client.get(reverse("transfer_list"))
        self.assertContains(transfers, self.a_transfer.code)
        self.assertNotContains(transfers, self.b_origin.name)
        self.assertNotContains(transfers, str(self.b_transfer.uuid))
        self.assertEqual(self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,))).status_code, 200)
        self.assertEqual(self.client.get(reverse("reports")).status_code, 200)
        self.assertEqual(self.client.get(reverse("audit_list")).status_code, 200)

        for route_name in ("manage_branches", "manage_products", "manage_users", "manage_assignments"):
            response = self.client.get(reverse(route_name))
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Solo lectura")
        users = self.client.get(reverse("manage_users"))
        self.assertContains(users, self.a_admin.username)
        self.assertNotContains(users, self.b_auditor.username)
        self.assertNotContains(users, "Nuevo usuario")

        original_status = self.a_transfer.status
        self.assertEqual(
            self.client.post(reverse("transfer_prepare", args=(self.a_transfer.uuid,))).status_code,
            403,
        )
        self.a_transfer.refresh_from_db()
        self.assertEqual(self.a_transfer.status, original_status)
        self.assertEqual(
            self.client.post(reverse("manage_products"), {
                "code": "FORBIDDEN",
                "name": "No debe crearse",
                "category": "Auditoría",
            }).status_code,
            403,
        )
        self.assertFalse(Product.objects.filter(company=self.company_a, code="FORBIDDEN").exists())
        self.assertEqual(
            self.client.get(reverse("manage_user_edit", args=(self.a_manager.pk,))).status_code,
            403,
        )
        self.assertEqual(
            self.client.post(reverse("commercial_register", args=(self.a_transfer.uuid,)), {
                "external_reference": "AUDIT-NO",
                "external_date": date.today().isoformat(),
                "notes": "",
                "correction_reason": "",
            }).status_code,
            403,
        )

    def test_auditor_cannot_discover_another_company(self):
        self.client.force_login(self.a_auditor)
        self.assertEqual(
            self.client.get(reverse("transfer_detail", args=(self.b_transfer.uuid,))).status_code,
            404,
        )
        self.assertEqual(
            self.client.get(reverse("evidence_download", args=(self.b_evidence.uuid,))).status_code,
            404,
        )
        self.assertEqual(
            self.client.post(reverse("commercial_register", args=(self.b_transfer.uuid,)), {
                "external_reference": "FORBIDDEN",
                "external_date": date.today().isoformat(),
                "notes": "",
                "correction_reason": "",
            }).status_code,
            404,
        )

    def test_reports_only_count_visible_company(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("reports"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total"], 1)
        self.assertNotContains(response, self.b_origin.name)

        future = (date.today() + timedelta(days=1)).isoformat()
        filtered = self.client.get(reverse("reports"), {"date_from": future})
        self.assertEqual(filtered.context["total"], 0)
        self.assertContains(filtered, f'value="{future}"')

    @override_settings(EVIDENCE_MAX_FILE_SIZE_MB=1)
    def test_evidence_upload_rejects_disallowed_type_oversize_and_false_signature(self):
        self.client.force_login(self.a_manager)
        upload_url = reverse("upload_evidence", args=(self.a_transfer.uuid,))

        invalid_type = self.client.post(upload_url, {
            "type": Evidence.Type.PREPARATION,
            "description": "GIF no permitido",
            "file": SimpleUploadedFile("evidence.gif", b"GIF89a-invalid", content_type="image/gif"),
        }, follow=True)
        self.assertContains(invalid_type, "Formato no permitido")

        oversized = self.client.post(upload_url, {
            "type": Evidence.Type.PREPARATION,
            "description": "Archivo grande",
            "file": SimpleUploadedFile(
                "large.jpg",
                b"\xff\xd8\xff" + (b"0" * (1024 * 1024)),
                content_type="image/jpeg",
            ),
        }, follow=True)
        self.assertContains(oversized, "supera el límite de 1 MB")

        mismatched_type = self.client.post(upload_url, {
            "type": Evidence.Type.PREPARATION,
            "description": "MIME incorrecto",
            "file": SimpleUploadedFile("photo.jpg", b"\xff\xd8\xffvalid-header", content_type="image/png"),
        }, follow=True)
        self.assertContains(mismatched_type, "tipo de archivo no coincide con su extensión")

        false_signature = self.client.post(upload_url, {
            "type": Evidence.Type.PREPARATION,
            "description": "Contenido falso",
            "file": SimpleUploadedFile("fake.png", b"%PDF-fake", content_type="image/png"),
        }, follow=True)
        self.assertContains(false_signature, "contenido del archivo no coincide")
        self.assertEqual(self.a_transfer.evidences.count(), 0)

        evidence_page = self.client.get(reverse("transfer_detail", args=(self.a_transfer.uuid,)))
        self.assertContains(evidence_page, "JPG, PNG, WEBP o PDF · máximo 1 MB.")
        self.assertContains(evidence_page, 'accept=".jpg,.jpeg,.png,.webp,.pdf,image/jpeg,image/png,image/webp,application/pdf"')

    def test_advanced_filter_cannot_expand_tenant_scope(self):
        self.client.force_login(self.a_admin)
        response = self.client.get(reverse("transfer_list"), {"product": self.b_product.pk})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["transfers"]), [])

    def test_notification_is_marked_read_when_opened(self):
        notification = Notification.objects.create(user=self.a_admin, transfer=self.a_transfer, message="Pendiente")
        self.client.force_login(self.a_admin)
        dashboard = self.client.get(reverse("dashboard"))
        self.assertContains(dashboard, 'class="indicator"')
        self.assertContains(dashboard, "badge badge-error badge-xs")
        self.assertNotContains(dashboard, "badge-error badge-outline badge-xs")
        response = self.client.get(reverse("notification_open", args=(notification.pk,)))
        self.assertRedirects(response, f"{reverse('transfer_detail', args=(self.a_transfer.uuid,))}?flow=1")
        notification.refresh_from_db()
        self.assertTrue(notification.is_read)

    def test_login_logout_and_password_change_work(self):
        self.assertTrue(self.client.login(username=self.a_admin.username, password="x"))
        self.assertTrue(AuditLog.objects.filter(user=self.a_admin, action="LOGIN").exists())
        response = self.client.post(reverse("password_change"), {
            "old_password": "x",
            "new_password1": "a-new-secure-password-9384",
            "new_password2": "a-new-secure-password-9384",
        })
        self.assertRedirects(response, reverse("dashboard"))
        self.a_admin.refresh_from_db()
        self.assertTrue(self.a_admin.check_password("a-new-secure-password-9384"))
        self.client.post(reverse("logout"))
        self.assertTrue(AuditLog.objects.filter(user=self.a_admin, action="LOGOUT").exists())


class SequenceConcurrencyTests(TransactionTestCase):
    reset_sequences = True

    def setUp(self):
        self.company = Company.objects.create(code="concurrent", name="Concurrente")

    def _take(self, _):
        close_old_connections()
        try:
            return Sequence.take(self.company, Sequence.Kind.TRANSFER, 2026)
        finally:
            close_old_connections()

    def test_sequence_values_are_unique_under_parallel_writes(self):
        with ThreadPoolExecutor(max_workers=4) as executor:
            values = list(executor.map(self._take, range(12)))
        self.assertEqual(sorted(values), list(range(1, 13)))


class SuperuserAdminTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.company = Company.objects.create(code="admin-scope", name="Empresa técnica")
        cls.branch = Branch.objects.create(company=cls.company, code="TEC", name="Sucursal técnica")
        cls.superuser = User.objects.create_superuser(
            username="root-technical",
            password="test-pass-123",
            email="root@example.com",
        )

    def setUp(self):
        self.client.force_login(self.superuser)

    def test_superuser_is_technical_only_and_redirected_to_admin(self):
        self.assertFalse(self.superuser.has_capability("prepare"))
        self.assertFalse(self.superuser.has_capability("reconcile"))
        response = self.client.get(reverse("dashboard"))
        self.assertRedirects(response, reverse("superadmin:index"), fetch_redirect_response=False)

    def test_admin_only_allows_creating_companies_and_company_admins(self):
        self.assertEqual(self.client.get(reverse("superadmin:core_company_add")).status_code, 200)
        user_add = self.client.get(reverse("superadmin:core_user_add"))
        self.assertEqual(user_add.status_code, 200)
        self.assertContains(user_add, "Django Admin solo crea administradores")
        self.assertNotContains(user_add, 'name="role"')
        self.assertNotContains(user_add, 'name="branch"')
        self.assertEqual(self.client.get(reverse("superadmin:core_branch_add")).status_code, 403)
        self.assertEqual(self.client.get(reverse("superadmin:core_product_add")).status_code, 403)
        self.assertEqual(self.client.get(reverse("superadmin:core_transfer_add")).status_code, 403)

    def test_admin_user_creation_always_creates_company_admin(self):
        response = self.client.post(reverse("superadmin:core_user_add"), {
            "username": "first-company-admin",
            "usable_password": "true",
            "password1": "Granite!9384-River",
            "password2": "Granite!9384-River",
            "company": self.company.pk,
            "first_name": "Ada",
            "last_name": "Flores",
            "email": "ada@example.com",
            "is_active": "on",
            "role": User.Role.MANAGER,
            "branch": self.branch.pk,
        })
        self.assertEqual(response.status_code, 302, getattr(response, "context", None))
        user = User.objects.get(username="first-company-admin")
        self.assertEqual(user.role, User.Role.COMPANY_ADMIN)
        self.assertEqual(user.company, self.company)
        self.assertIsNone(user.branch_id)
        self.assertFalse(user.is_staff)
        self.assertFalse(user.is_superuser)

    def test_existing_data_correction_requires_reason_and_is_audited(self):
        change_url = reverse("superadmin:core_company_change", args=(self.company.pk,))
        without_reason = self.client.post(change_url, {
            "code": self.company.code,
            "name": "Nombre sin justificar",
            "is_active": "on",
            "correction_reason": "",
        })
        self.assertEqual(without_reason.status_code, 200)
        self.assertContains(without_reason, "Indica por qué se realiza esta corrección")
        self.company.refresh_from_db()
        self.assertEqual(self.company.name, "Empresa técnica")

        with_reason = self.client.post(change_url, {
            "code": self.company.code,
            "name": "Nombre corregido",
            "is_active": "on",
            "correction_reason": "Corrección solicitada por soporte con respaldo interno.",
        })
        self.assertEqual(with_reason.status_code, 302)
        self.company.refresh_from_db()
        self.assertEqual(self.company.name, "Nombre corregido")
        log = AuditLog.objects.get(action="SUPERADMIN_UPDATE", object_uuid=self.company.code)
        self.assertEqual(log.reason, "Corrección solicitada por soporte con respaldo interno.")
        self.assertEqual(log.before["name"], "Empresa técnica")
        self.assertEqual(log.after["name"], "Nombre corregido")
